"""生成 template_panel.xlsx —— 100 股面板快照模板（可复现）。

关键经验（踩坑记录）：
1. 基于 template_v2.xlsx 生成，继承 Wind 的 externalLinks（[1]! 指向 windfunc.xlam）
2. 纯 openpyxl 生成 → Excel 不报"修复"（zipfile 粗暴重打包会破坏文件完整性）
3. openpyxl save 后 externalBook r:id 与 rels Id 不匹配（rId1 vs rId3）→ [1]! 失效
4. 用"保守 zipfile"（保留每个成员的 ZipInfo + 压缩方式 + 顺序）只改 rels 一处对齐 r:id
   → 既不破坏文件、又让 [1]! 工作

用法：
    uv run python scripts/make_panel_template.py
"""
import csv
import zipfile
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent.parent.resolve()
SRC = ROOT / "template_v2.xlsx"                       # 干净源（含 externalLinks）
OUT = ROOT / "templates" / "template_panel.xlsx"
CODES = ROOT / "config" / "codes.csv"

HF = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="2878B5")


def read_codes():
    rows = []
    with open(CODES, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append((r["wind_code"].strip(), r.get("name", "").strip()))
    return rows


def build(rows):
    n = len(rows)
    wb = load_workbook(SRC)
    for name in list(wb.sheetnames):
        if name in ("Sheet1", "consensus_fy1", "daily_price_vol", "wind_code", "benchmark") \
                or name.strip() == "static_info":
            del wb[name]

    # _meta
    ws = wb.create_sheet("_meta", 0)
    ws.append(["wind_code", "name"])
    for c, nm in rows:
        ws.append([c, nm])
    for c in range(1, 3):
        ws.cell(row=1, column=c).font = HF
        ws.cell(row=1, column=c).fill = HFILL

    # panel
    ws = wb.create_sheet("panel", 1)
    ws["A1"] = "观测日期→"; ws["B1"] = "=TODAY()"; ws["A1"].font = Font(bold=True)
    hdr = ["代码", "日期", "FY1净利润均值", "FY1_EPS", "FY1机构数", "FY1标准差", "FY1中值",
           "FY2净利润均值", "FY2_EPS", "FY2机构数", "FY2标准差", "FY2中值",
           "收盘价", "成交量", "成交额", "换手率", "滚动PE", "总市值(亿)"]
    for i, h in enumerate(hdr, 1):
        cc = ws.cell(row=2, column=i, value=h)
        cc.font = HF; cc.fill = HFILL; cc.alignment = Alignment(horizontal="center")
    for k in range(n):
        r = 3 + k; A = f"$A{r}"; B = f"$B{r}"
        ws.cell(row=r, column=1, value=f"=_meta!A{2+k}")
        ws.cell(row=r, column=2, value="=$B$1")
        ws.cell(row=r, column=3,  value=f"=[1]!s_west_netprofit({A},YEAR({B}),{B},180,1000000)")
        ws.cell(row=r, column=4,  value=f"=[1]!s_west_eps({A},YEAR({B}),{B},180)")
        ws.cell(row=r, column=5,  value=f"=[1]!s_west_instnum_np({A},YEAR({B}),{B},180)")
        ws.cell(row=r, column=6,  value=f"=[1]!s_west_stdnetprofit({A},YEAR({B}),{B},180,1000000)")
        ws.cell(row=r, column=7,  value=f"=[1]!s_west_mediannetprofit({A},YEAR({B}),{B},180,1000000)")
        ws.cell(row=r, column=8,  value=f"=[1]!s_west_netprofit({A},YEAR({B})+1,{B},180,1000000)")
        ws.cell(row=r, column=9,  value=f"=[1]!s_west_eps({A},YEAR({B})+1,{B},180)")
        ws.cell(row=r, column=10, value=f"=[1]!s_west_instnum_np({A},YEAR({B})+1,{B},180)")
        ws.cell(row=r, column=11, value=f"=[1]!s_west_stdnetprofit({A},YEAR({B})+1,{B},180,1000000)")
        ws.cell(row=r, column=12, value=f"=[1]!s_west_mediannetprofit({A},YEAR({B})+1,{B},180,1000000)")
        ws.cell(row=r, column=13, value=f"=[1]!s_dq_close({A},{B},3)")
        ws.cell(row=r, column=14, value=f"=[1]!s_dq_volume({A},{B})")
        ws.cell(row=r, column=15, value=f"=[1]!s_dq_amount({A},{B})")
        ws.cell(row=r, column=16, value=f"=[1]!s_dq_turn({A},{B})")
        ws.cell(row=r, column=17, value=f"=[1]!s_val_pe_ttm({A},{B})")
        ws.cell(row=r, column=18, value=f"=[1]!S_VAL_MV({A},{B},100000000)")
    ws.freeze_panes = "C3"

    # static_info
    ws = wb.create_sheet("static_info")
    ws["A1"] = "基准日→"; ws["B1"] = "=TODAY()"; ws["A1"].font = Font(bold=True)
    for i, h in enumerate(["代码", "公司名", "申万一级", "申万二级", "上市日期"], 1):
        cc = ws.cell(row=2, column=i, value=h); cc.font = HF; cc.fill = HFILL
    for k in range(n):
        r = 3 + k; A = f"$A{r}"
        ws.cell(row=r, column=1, value=f"=_meta!A{2+k}")
        ws.cell(row=r, column=2, value=f"=[1]!S_INFO_NAME({A})")
        ws.cell(row=r, column=3, value=f"=[1]!hks_info_industry_sw_2021({A},$B$1,1)")
        ws.cell(row=r, column=4, value=f"=[1]!hks_info_industry_sw_2021({A},$B$1,2)")
        ws.cell(row=r, column=5, value=f"=[1]!s_ipo_listeddate({A})")

    # benchmark
    ws = wb.create_sheet("benchmark")
    ws["A1"] = "观测日期→"; ws["B1"] = "=TODAY()"; ws["A1"].font = Font(bold=True)
    for i, h in enumerate(["代码", "收盘", "名称"], 1):
        cc = ws.cell(row=2, column=i, value=h); cc.font = HF; cc.fill = HFILL
    ws["A3"] = "HSTECH.HI"; ws["B3"] = '=[1]!s_dq_close($A3,$B$1,3)'; ws["C3"] = "恒生科技指数"
    ws["A4"] = "HSI.HI";    ws["B4"] = '=[1]!s_dq_close($A4,$B$1,3)'; ws["C4"] = "恒生指数"

    wb.save(OUT)


def fix_relid():
    """保守 zipfile：对齐 externalBook r:id 与 rels Id，保留 ZipInfo 不破坏文件。"""
    tmp = str(OUT) + ".fix"
    target = "xl/externalLinks/_rels/externalLink1.xml.rels"
    with zipfile.ZipFile(OUT, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
        for item in zin.infolist():                       # 保留原顺序
            data = zin.read(item.filename)
            if item.filename == target:
                data = data.replace(b'rId3', b'rId1').replace(b'rId2', b'rId1')
            zout.writestr(item, data, compress_type=item.compress_type)  # 保留压缩方式
    shutil.move(tmp, OUT)


def main():
    rows = read_codes()
    build(rows)
    fix_relid()
    # 验证
    import re
    with zipfile.ZipFile(OUT) as z:
        eb = z.read("xl/externalLinks/externalLink1.xml").decode()
        rels = z.read("xl/externalLinks/_rels/externalLink1.xml.rels").decode()
    rid_eb = re.search(r'r:id="([^"]+)"', eb).group(1)
    rid_rels = re.search(r'Id="([^"]+)"', rels).group(1)
    sheets = load_workbook(OUT, read_only=True).sheetnames
    print(f"✅ 生成 {OUT}")
    print(f"   sheets: {sheets}")
    print(f"   股票数: {len(rows)} | r:id 匹配: {rid_eb == rid_rels} ({rid_eb}={rid_rels})")


if __name__ == "__main__":
    main()
