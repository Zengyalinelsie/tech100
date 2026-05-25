"""生成 templates/template_universe_daily.xlsx — V8 日频 universe 采集模板。

布局:
    B1                    = 观测日期(字符串 YYYY-MM-DD,采集脚本动态写)
    A1                    = "观测日期→"(说明文字)
    A2:G2                 = 表头(代码 / 日期 / 收盘 / PE / PB_lf / 股息率 / FY1 EPS)
    A3:A35                = 33 个 universe 代码(从 db.universe 表读)
    B3:B35                = =$B$1  (每行日期 = B1)
    C3:G35                = Wind 公式

字段(P0 验证通过):
    C: 收盘点位      =[1]!s_dq_close($A3,$B3)
    D: PE_TTM       =[1]!s_val_pe_ttm($A3,$B3)
    E: PB_LF        =[1]!s_val_pb_lf($A3,$B3)
    F: 股息率 %     =[1]!s_val_dividendyield2($A3,$B3)
    G: FY1 EPS      =[1]!s_west_eps($A3,YEAR($B3),$B3,180)

用法:
    uv run python scripts/make_universe_template.py
"""
import os
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(ROOT, "data", "wind_history.db")
OUT = os.path.join(ROOT, "templates", "template_universe_daily.xlsx")

FIELDS = [
    ("收盘",       '=[1]!s_dq_close($A{r},$B{r})'),
    ("PE_TTM",     '=[1]!s_val_pe_ttm($A{r},$B{r})'),
    ("PB_LF",      '=[1]!s_val_pb_lf($A{r},$B{r})'),
    ("股息率%",    '=[1]!s_val_dividendyield2($A{r},$B{r})'),
    ("FY1_EPS",    '=[1]!s_west_eps($A{r},YEAR($B{r}),$B{r},180)'),
]


def main():
    conn = sqlite3.connect(DB_PATH)
    rows = list(conn.execute(
        "SELECT universe_code, name_cn, market, universe_type FROM universe "
        "ORDER BY market, universe_type, universe_code"
    ))
    conn.close()
    if not rows:
        raise SystemExit("universe 表空,先跑 seed_universe.py")

    wb = Workbook()
    ws = wb.active
    ws.title = "universe"

    # 第 1 行:B1 日期参数
    ws["A1"] = "观测日期→"
    ws["A1"].font = Font(bold=True, color="666666")
    ws["B1"] = "2026-05-25"     # 占位字符串,采集脚本运行时覆盖
    ws["B1"].font = Font(bold=True, size=12)
    ws["B1"].fill = PatternFill("solid", fgColor="FFF2CC")
    # 注释列(第 1 行)
    ws["I1"] = "字段说明 → C 收盘 / D PE_TTM / E PB_lf(最近报告期) / F 股息率% / G FY1 EPS 一致预期(180 天窗口)"
    ws["I1"].font = Font(italic=True, color="999999", size=9)

    # 第 2 行:表头
    headers = ["代码", "日期", "收盘", "PE_TTM", "PB_LF", "股息率%", "FY1_EPS", "名称(参考)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    # 第 3 行起:33 universe × 5 字段公式 + H 列名称参考
    for i, (code, name, market, utype) in enumerate(rows):
        r = i + 3                                # 数据起 A3
        ws.cell(row=r, column=1, value=code)     # A
        ws.cell(row=r, column=2, value=f"=$B$1") # B 日期引用
        for j, (_, formula_tpl) in enumerate(FIELDS):
            ws.cell(row=r, column=3 + j, value=formula_tpl.format(r=r))   # C-G
        ws.cell(row=r, column=8, value=f"{name} ({market}/{utype})")      # H 参考名

    # 列宽
    widths = [14, 12, 10, 10, 10, 10, 10, 35]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"生成 {OUT}")
    print(f"  {len(rows)} 个 universe × 5 字段公式 = {len(rows) * 5} 个 Wind 公式")
    print(f"  A 股: {sum(1 for r in rows if r[2]=='A')} | 港股: {sum(1 for r in rows if r[2]=='HK')}")


if __name__ == "__main__":
    main()
